import sys
from PyQt6.QtWidgets import *
from PyQt6.QtGui import QPixmap, QCloseEvent
from PyQt6.QtCore import Qt
import mysql.connector
import openpyxl
from openpyxl.styles import Alignment


class UserInterface(QWidget):
    def __init__(self):
        super().__init__()
        self.transactions = []

        self.setGeometry(730, 300, 500, 500)
        self.connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="Transactions"
        )

        self.cursor = self.connection.cursor()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Интерфейс пользователя')

        layout = QVBoxLayout()

        # Просмотреть транзакции
        view_transactions_layout = QVBoxLayout()
        self.transactions_list = QListWidget()
        view_transactions_layout.addWidget(QLabel('История транзакций:'))
        view_transactions_layout.addWidget(self.transactions_list)
        export_button = QPushButton('Экспорт в Excel', self)
        export_button.clicked.connect(self.export_to_excel)

        layout.addLayout(view_transactions_layout)
        layout.addWidget(export_button)

        self.setLayout(layout)

        self.fetch_transactions_from_database()

    def fetch_transactions_from_database(self):
        try:
            # Выполните SQL-запрос для выборки транзакций из базы данных
            select_query = "SELECT buyer, seller, car_info, date, Address, Cost FROM transaction"
            self.cursor.execute(select_query)

            # Получите все строки результата
            rows = self.cursor.fetchall()

            # Преобразуйте каждую строку в объект Transaction и добавьте в список
            for row in rows:
                buyer, seller, car_info, date, place_of_contract, car_cost = row
                transaction = Transaction(buyer, seller, car_info, date, place_of_contract, car_cost)
                self.transactions.append(transaction)

            self.update_transactions_list()

        except mysql.connector.Error as err:
            print(f"Error: {err}")
            QMessageBox.warning(self, 'Error', f'Error fetching transactions from the database: {err}')

    def update_transactions_list(self):
        self.transactions_list.clear()
        for transaction in self.transactions:
            self.transactions_list.addItem(f'Покупатель: {transaction.buyer}, '
                                           f'Продавец: {transaction.seller}, '
                                           f'Марка: {transaction.car_info}, '
                                           f'Дата: {transaction.date}, '
                                           f'Место подписания договора: {transaction.place_of_contract}, '
                                           f'Стоимость: {transaction.car_cost}')

    def export_to_excel(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            headers = ["Покупатель", "Продавец", "Марка авто", "Дата транзакции", "Место подписания договора",
                       "Стоимость"]
            sheet.append(headers)

            for transaction in self.transactions:
                row_data = [
                    transaction.buyer,
                    transaction.seller,
                    transaction.car_info,
                    transaction.date,
                    transaction.place_of_contract,
                    transaction.car_cost
                ]
                sheet.append(row_data)

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

            save_path, _ = QFileDialog.getSaveFileName(self, 'Сохранить в Excel', '', 'Excel Files (*.xlsx)')
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

            if save_path:
                workbook.save(save_path)

                QMessageBox.information(self, 'Экспорт в Excel', 'Данные успешно экспортированы в Excel.')

        except Exception as e:
            print(f"Произошла ошибка: {e}")
            QMessageBox.critical(self, 'Ошибка', f'Произошла ошибка при экспорте в Excel: {e}')


class User:
    def __init__(self, username, password, role):
        self.username = username
        self.password = password
        self.role = role


class Transaction:
    def __init__(self, buyer='', seller='', car_info='', date='', place_of_contract='', car_cost=''):
        self.buyer = buyer
        self.seller = seller
        self.car_info = car_info
        self.date = date
        self.place_of_contract = place_of_contract
        self.car_cost = car_cost
